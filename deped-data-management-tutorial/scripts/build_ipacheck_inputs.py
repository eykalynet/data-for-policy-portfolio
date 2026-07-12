from openpyxl import Workbook

ppi_vars = [
    "ppi_q1_floor_solid",
    "ppi_q2_has_toilet",
    "ppi_q3_has_electricity",
    "ppi_q4_has_tv",
    "ppi_q5_has_fridge",
    "ppi_q6_head_completed_primary",
    "ppi_q7_roof_durable",
    "ppi_q8_has_mobile_phone",
    "ppi_q9_owns_livestock",
    "ppi_q10_has_savings",
]

wb = Workbook()
ws = wb.active
ws.title = "constraints"
ws.append(["variable", "label", "hard_min", "soft_min", "soft_max", "hard_max", "notes", "keep"])

for var in ppi_vars:
    ws.append([var, "PPI binary item should be 0 or 1", 0, "", "", 1, "", "key barangay form_version"])

ws.append(["duration_minutes", "Survey duration should be plausible", 8, "", "", 90, "", "key barangay form_version"])
ws.append(["hh_size", "Household size should be plausible", 1, "", "", 15, "", "key barangay form_version"])
ws.append(["children_under_15", "Children under 15 should be plausible", 0, "", "", 15, "", "key barangay form_version"])

ws = wb.create_sheet("logic")
ws.append(["variable", "label", "assert", "if_condition", "notes", "keep"])
ws.append(["children_under_15", "Children cannot exceed household size", "children_under_15 <= hh_size", "", "", "key barangay hh_size"])

ws = wb.create_sheet("outliers")
ws.append(["variable", "label", "by", "method", "multiplier", "combine", "notes", "keep"])
ws.append(["duration_minutes", "Survey duration", "enumerator", "iqr", 1.5, "", "", "key barangay form_version"])

ws = wb.create_sheet("other specify")
ws.append(["parent", "parent label", "child", "child label", "notes", "keep"])
ws.append([
    "asset_main",
    "Main asset selected",
    "asset_other",
    "Other asset specified",
    "Lists open-text values when respondent selected other",
    "key barangay",
])

ws = wb.create_sheet("enumstats")
ws.append(["variable", "label", "min", "mean", "show_mean_as", "median", "sd", "max", "combine", "notes"])
ws.append(["duration_minutes", "Survey duration", "yes", "yes", "number", "yes", "yes", "yes", "", "Monitor speed by enumerator"])
ws.append(["ppi_score", "PPI score", "yes", "yes", "number", "yes", "yes", "yes", "", "Monitor PPI score distribution by enumerator"])

wb.save("inputs/ipacheck_inputs.xlsx")
print("Created inputs/ipacheck_inputs.xlsx")
